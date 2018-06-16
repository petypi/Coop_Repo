"""
Main library for processing the deliveries intended for customers
with more than one confirmed orders
"""
import psycopg2
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders as Encoders
import ast
import smtplib
import ConfigParser
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Color

CONFIG_READER = ConfigParser.ConfigParser()
CONFIG_READER.read(os.path.join(os.path.dirname(__file__), 'settings.ini'))
EMAIL_SETTINGS = dict(CONFIG_READER.items('email'))
APP_SETTINGS = dict(CONFIG_READER.items('data'))
DB_SETTINGS = dict(CONFIG_READER.items('database'))


class Orders(object):
    """
    Fetch customers with multiple orders from the database and mail them to concerned
    party
    """
    def __init__(self, db_connection):
        self.db_management = DatabaseManagement(db_connection)

    def email_results(self, filename):
        """
        Send the files to the team involved
        """
        try:
            message = MIMEMultipart('alternative')
            message['From'] = "Copia Report Automation"
            toaddr = ast.literal_eval(EMAIL_SETTINGS['recipients'])
            message['Subject'] = EMAIL_SETTINGS['subject']
            message.attach(MIMEText("Dear Team,\nAttached is today's Customers with multiple confirmed orders report."))
            attachment = MIMEBase('application', 'octet-stream')
            attachment.set_payload(file(filename).read())
            Encoders.encode_base64(attachment)
            attachment.add_header('Content-Disposition',
                                  'attachment',
                                  filename=filename)
            message.attach(attachment)
            password = EMAIL_SETTINGS['password']
            server = smtplib.SMTP(EMAIL_SETTINGS['server'],
                                  int(EMAIL_SETTINGS['port']))
            server.ehlo()
            server.starttls()
            server.login(EMAIL_SETTINGS['username'], password)
            server.sendmail(EMAIL_SETTINGS['username'], toaddr,
                            message.as_string())
            server.close()
        except smtplib.SMTPException:
            raise

    def get_new_agent_orders(self):
        """
        Get customers with multiple confirmed orders
        """
        order = None
        today = datetime.today().date()

        try:
            sql = '''
            select rp.name as customername, rp.phone as customerphone, rp2.name as agentname, 
            rp2.phone as agentphone, rp3.name as associatename, so.name as saleorderid,
            pt.default_code, pt.name, pt.list_price, sol.product_uom_qty, so.amount_total, 
            date_order, date_delivery from sale_order so
            left join res_partner rp on rp.id = so.customer_id
            left join res_partner rp2 on so.partner_id = rp2.id
            left join res_partner rp3 on rp2.sale_associate_id = rp3.id
            left Join sale_order_line sol on so.id = sol.order_id
            left Join product_product pp on sol.product_id=pp.id
            left Join product_template pt on pp.product_tmpl_id=pt.id
            where confirmation_date::date = %s and so.customer_id in 
            (Select customer_id from sale_order Where confirmation_date = %s
            Group By customer_id Having Count(customer_id) > 1)Order by so.customer_id;
            '''
            # params = (str(today), )
            params = (str(today), str(today))
            order = self.db_management.retrieve_all_data_params(sql, params)
        except psycopg2.Error:
            raise
        return order

    def format_xlsx(self, orders):
        """
        Turn the received dictionary into an xlsx file to be sent out
        """
        result_file = Workbook()

        active_sheet = result_file.active
        headers = ['Customer Name', 'Customer Number', 'Agent Name', 'Agent Number', 'Associate Name', 'Order ID',
                   'Product Code', 'Product Description', 'Product Price', 'Product Quantity', 'Order Total',
                   'Date Order', 'Delivery Date']
        empty_field = ['', '', '', '', '', '', '', '', '', '', '', '', '', ]

        row_num = len(headers)
        # cel_merge = 'A1:K1'
        # active_sheet.merge_cells(cel_merge)
        active_sheet.append(['Multiple Orders Customers'])
        active_sheet.append(empty_field)
        # active_sheet.font = Font(size=22)
        active_sheet.append(headers)
        # count = 3
        for order in orders:
            my_order = ["'" + str(item)
                        if isinstance(item, datetime)
                        else item
                        for item in order]
            active_sheet.append(my_order)
            # count = count + 1
        # cel_merge = 'A%s:K%s' % (str(count), str(count))
        active_sheet.append(empty_field)
        active_sheet.append(empty_field)
        # active_sheet.merge_cells(cel_merge)

        result_file.save('multiple_orders_customers.xlsx')


class DatabaseManagement(object):
    """
    The basic class for DB management methods
    """
    def __init__(self, db_connection):
        self.db_connection = db_connection

    def retrieve_all_data_params(self, sql, params):
        """
        Retrieve all data associated with the params from the database
        """
        results = None
        try:
            cursor = self.db_connection.cursor()
            cursor.execute(sql, params)
            results = cursor.fetchall()
        except psycopg2.Error:
            raise
        return results

    def execute_query(self, sql, params):
        """
        Run a query that does not require a return
        """
        try:
            cursor = self.db_connection.cursor()
            cursor.execute(sql, params)
            self.db_connection.commit()
        except psycopg2.DatabaseError:
            raise
