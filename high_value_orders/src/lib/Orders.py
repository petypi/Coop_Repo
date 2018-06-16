"""
Main library for processing the deliveries intended for new agents
"""
import psycopg2
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders as Encoders
import ast
import smtplib
import ConfigParser
import os
from openpyxl import Workbook
# from openpyxl.styles import Font, Color

CONFIG_READER = ConfigParser.ConfigParser()
CONFIG_READER.read(os.path.join(os.path.dirname(__file__), 'settings.ini'))
EMAIL_SETTINGS = dict(CONFIG_READER.items('email'))
APP_SETTINGS = dict(CONFIG_READER.items('data'))
DB_SETTINGS = dict(CONFIG_READER.items('database'))


class Orders(object):
    """
    Fetch newest product changes from the database and mail them to concerned
    party
    """
    def __init__(self, db_connection):
        self.db_management = DatabaseManagement(db_connection)

    def email_results(self, filename):
        """
        Send the files to the team involved
        """
        try:
            message = MIMEMultipart('alternative')
            message['From'] = "Copia Report Automation"
            toaddr = ast.literal_eval(EMAIL_SETTINGS['recipients'])
            message['Subject'] = EMAIL_SETTINGS['subject']
            message.attach(MIMEText("Dear Team,\nAttached is today's High Value Orders report."))
            attachment = MIMEBase('application', 'octet-stream')
            attachment.set_payload(file(filename).read())
            Encoders.encode_base64(attachment)
            attachment.add_header('Content-Disposition',
                                  'attachment',
                                  filename=filename)
            message.attach(attachment)
            password = EMAIL_SETTINGS['password']
            server = smtplib.SMTP(EMAIL_SETTINGS['server'],
                                  int(EMAIL_SETTINGS['port']))
            server.ehlo()
            server.starttls()
            server.login(EMAIL_SETTINGS['username'], password)
            server.sendmail(EMAIL_SETTINGS['username'], toaddr,
                            message.as_string())
            server.close()
        except smtplib.SMTPException:
            raise

    def get_new_agent_orders(self, state='draft'):
        """
        Get the orders with total price > 10000
        """
        order = None
        jana = datetime.today().date() - timedelta(days=1)
        if datetime.weekday(jana) == 6:
            jana = datetime.today().date() - timedelta(days=2)
        try:
            sql = ('''
                    SELECT
                        so.date_delivery,so.name, pp.default_code, pt.name,
                        pt.list_price, sol.product_uom_qty, so.amount_total, rp.name,
                        rp.phone, rp1.phone, sa.name, dr.name
                    FROM sale_order so
                    LEFT JOIN sale_order_line sol
                        ON so.id= sol.order_id
                    LEFT JOIN product_product pp
                        ON sol.product_id=pp.id
                    LEFT JOIN product_template pt
                        ON pp.product_tmpl_id =pt.id
                    LEFT JOIN res_partner rp
                        ON so.partner_id=rp.id
                    LEFT JOIN res_partner rp1
                        ON rp1.id = so.customer_id
                    LEFT JOIN res_partner sa
                        ON sa.id = so.sale_associate_id
                    LEFT JOIN res_partner_data rpd
                        ON rpd.partner_id = rp.id
                    LEFT JOIN delivery_route dr
                        ON dr.id = rpd.route_id
                    WHERE so.amount_total > 10000
                    AND so.state = %s
                    AND so.date_order = %s;
                    '''
                   )
            params = (state, str(jana),)
            order = self.db_management.retrieve_all_data_params(sql, params)
        except psycopg2.Error:
            raise
        return order

    def get_high_value_product(self, state='draft'):
        """
        get products with price greater than 6000
        """
        product = None
        jana = datetime.today().date() - timedelta(days=1)
        try:
            sql = ('''
                    SELECT
                        so.date_delivery,so.name, pp.default_code, pt.name,
                        pt.list_price, sol.product_uom_qty, so.amount_total, rp.name,
                        rp.phone, rp1.phone, sa.name, dr.name
                    FROM sale_order so
                    LEFT JOIN sale_order_line sol
                        ON so.id= sol.order_id
                    LEFT JOIN product_product pp
                        ON sol.product_id=pp.id
                    LEFT JOIN product_template pt
                        ON pp.product_tmpl_id =pt.id
                    LEFT JOIN res_partner rp
                        ON so.partner_id=rp.id
                    LEFT JOIN res_partner rp1
                        ON rp1.id = so.customer_id
                    LEFT JOIN res_partner sa
                        ON sa.id = so.sale_associate_id
                    LEFT JOIN res_partner_data rpd
                        ON rpd.partner_id = rp.id
                    LEFT JOIN delivery_route dr
                        ON dr.id = rpd.route_id
                    WHERE pt.list_price > 6000
                    AND so.state = %s
                    AND so.date_order = %s;
                    '''
                   )
            params = (state, str(jana),)
            product = self.db_management.retrieve_all_data_params(sql, params)
        except psycopg2.Error:
            raise
        return product

    def format_xlsx(self, orders, products, file_name):
        """
        Turn the received dictionary into an xlsx file to be sent out
        """
        result_file = Workbook()

        active_sheet = result_file.active
        headers = ['Date of Delivery', 'Order Number', 'Product Code',
                   'Product Description', 'Product Price', 'Product Quantity',
                   'Order Total', 'Agent Name', 'Agent Number', 'Customer Number',
                   'Sales Associate', 'Route Name']
        empty_field = ['', '', '', '', '', '','', '', '', '','']
        row_num = len(headers)
        # cel_merge = 'A1:K1'
        # active_sheet.merge_cells(cel_merge)
        active_sheet.append(['High Value Orders'])
        active_sheet.append(empty_field)
        # active_sheet.font = Font(size=22)
        active_sheet.append(headers)
        # count = 3
        for order in orders:
            my_order = ["'" + str(item)
                        if isinstance(item, datetime)
                        else item
                        for item in order]
            active_sheet.append(my_order)
            # count = count + 1
        # cel_merge = 'A%s:K%s' % (str(count), str(count))
        active_sheet.append(empty_field)
        active_sheet.append(empty_field)
        # active_sheet.merge_cells(cel_merge)
        active_sheet.append(['High Value Products'])
        active_sheet.append(empty_field)
        active_sheet.append(headers)
        for product in products:
            my_prod = ["'" + str(item)
                       if isinstance(item, datetime) == True
                       else item
                       for item in product]
            active_sheet.append(my_prod)

        result_file.save(file_name)


class DatabaseManagement(object):
    """
    The basic class for DB management methods
    """
    def __init__(self, db_connection):
        self.db_connection = db_connection

    def retrieve_all_data_params(self, sql, params):
        """
        Retrieve all data associated with the params from the database
        """
        results = None
        try:
            cursor = self.db_connection.cursor()
            cursor.execute(sql, params)
            results = cursor.fetchall()
        except psycopg2.Error:
            raise
        return results

    def execute_query(self, sql, params):
        """
        Run a query that does not require a return
        """
        try:
            cursor = self.db_connection.cursor()
            cursor.execute(sql, params)
            self.db_connection.commit()
        except psycopg2.DatabaseError:
            raise
